import itertools
import openpyxl
import sys
from pathlib import Path

from parser import parse, execute_all
from excel_helpers import parse_dinh_dang


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

EXCEL_PATH = Path(r"C:\Users\PhuThaiCAT\Desktop\Code\vnsky\dinh_dang_095_soluong.xlsx")


def generate_count(template: str, rule_str: str) -> int:
    conditions = parse(rule_str)
    if not conditions:
        return 0

    vars_ = []
    seen = set()
    for ch in template:
        if ch.isupper() and ch not in seen:
            seen.add(ch)
            vars_.append(ch)

    count = 0
    for combo in itertools.product("123456789", repeat=len(vars_)):
        ctx = dict(zip(vars_, combo))
        if all(execute_all(conditions, ctx)):
            count += 1
    return count


def process_all():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Định dạng 095"]
    ws.cell(row=1, column=7, value="Count")
    total = 0
    for row_idx in range(2, ws.max_row + 1):
        cell_val = str(ws.cell(row=row_idx, column=2).value or "")
        if cell_val is None:
            continue
        p = parse_dinh_dang(cell_val)
        req = p.get("requirements")
        if req:
            cnt = generate_count(p["number"], req)
        else:
            cnt = 0
        ws.cell(row=row_idx, column=7, value=cnt)
        total += 1
        print(f"  Row {row_idx}: {cnt}")
    wb.save(EXCEL_PATH)
    wb.close()
    print(f"\nDone — {total} rows processed.")


if __name__ == "__main__":
    process_all()
