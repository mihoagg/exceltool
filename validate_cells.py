import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

from excel_helpers import parse_dinh_dang
from parser import parse

load_dotenv()

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

EXCEL_PATH = Path(r"C:\Users\PhuThaiCAT\Desktop\Code\vnsky\excel\dinh_dang_095_soluong.xlsx")
SHEET_NAME = "Định dạng 095"
_API_BASES = {
    "groq": "https://api.groq.com/openai/v1/chat/completions",
    "openrouter": "https://openrouter.ai/api/v1/chat/completions",
}

def _detect_provider() -> str:
    if os.getenv("GROQ_API_KEY"):
        return "groq"
    if os.getenv("OPENROUTER_API_KEY"):
        return "openrouter"
    return "groq"

PROVIDER = _detect_provider()
API_BASE = _API_BASES[PROVIDER]
FAIL_LOG = Path("failures.jsonl")
BATCH_SIZE = 10
BATCH_DELAY = 1


def check_format(cell_value: str, parsed: dict) -> list[dict]:
    issues = []
    if not cell_value or not cell_value.strip():
        issues.append({"stage": "format", "type": "empty_cell", "severity": "fail", "detail": "Cell is empty"})
        return issues

    template = parsed.get("number", "")
    if not template:
        issues.append({"stage": "format", "type": "no_template", "severity": "fail", "detail": "Could not extract template from cell"})
        return issues

    paren_open = cell_value.count("(")
    paren_close = cell_value.count(")")
    if paren_open != paren_close:
        issues.append({
            "stage": "format", "type": "unbalanced_parens", "severity": "fail",
            "detail": f"Unbalanced parentheses: {paren_open} open vs {paren_close} close"
        })

    if paren_open > 0 and parsed.get("requirements") is None:
        issues.append({
            "stage": "format", "type": "unparseable_parens", "severity": "fail",
            "detail": "Cell contains parentheses but requirements could not be extracted"
        })

    return issues


def check_parse(requirements: str, conditions: list[dict]) -> list[dict]:
    issues = []
    if not requirements:
        return issues

    if not conditions:
        issues.append({
            "stage": "parse", "type": "empty_parse", "severity": "fail",
            "detail": f"Parser returned no conditions for: {requirements}"
        })
        return issues

    for i, cond in enumerate(conditions):
        if cond.get("type") == "raw":
            issues.append({
                "stage": "parse", "type": "raw_node", "severity": "warn",
                "detail": f"Condition {i + 1} is unparseable: {cond.get('raw', '?')}"
            })

    return issues


def check_variables(template: str, conditions: list[dict]) -> list[dict]:
    issues = []
    template_vars = set(ch for ch in template if ch.isupper() and ch.isalpha())
    if not template_vars:
        return issues

    flat = _flatten_conditions(conditions)
    rule_vars = set()
    for c in flat:
        f = c.get("field", "")
        if f.isalpha() and len(f) == 1:
            rule_vars.add(f)
    for v in sorted(rule_vars):
        if v not in template_vars:
            issues.append({
                "stage": "variable", "type": "missing_in_template", "severity": "fail",
                "detail": f"Variable '{v}' used in rules but not found in template '{template}'"
            })

    return issues


def _cond_to_str(cond: dict) -> str:
    t = cond.get("type")
    if t == "and":
        return "(" + " and ".join(_cond_to_str(c) for c in cond.get("children", [])) + ")"
    if t == "or":
        return "(" + " or ".join(_cond_to_str(c) for c in cond.get("children", [])) + ")"
    if t == "not":
        return "not " + _cond_to_str(cond["children"][0])
    field = cond.get("field", "?")
    vals = cond.get("values", [])
    if t == "equals":
        return f"{field} \u2208 {{{', '.join(vals)}}}"
    if t == "not_equals":
        return f"{field} \u2209 {{{', '.join(vals)}}}"
    op_map = {"gte": ">=", "lte": "<=", "gt": ">", "lt": "<"}
    op = op_map.get(t, "?")
    return f"{field} {op} {', '.join(vals)}"


def _flatten_conditions(conditions: list[dict]) -> list[dict]:
    flat = []
    for c in conditions:
        t = c.get("type")
        if t == "or":
            flat.append(c)
        elif t in ("and", "not"):
            flat.extend(_flatten_conditions(c.get("children", [])))
        else:
            flat.append(c)
    return flat


BATCH_SYSTEM_PROMPT = """
You validate a phone-number rule parser.

Each cell contains:
- a template
- parser output

The template is written for human interpretation, difficult to parse systematically. Validate ONLY whether the parser output is logically equivalent to the constraints implied by the template after applying the interpretation rules below. Ignore differences in formatting, ordering, grouping, or equivalent logical expressions.


Grammar

Variables:
A-Z represent digit positions.

Operators:
=  allowed values
#  not equal (displayed as ∉)
>, >=, <, <=

Rules:
A=1,2,B      -> A ∈ {1,2,B}
A#B          -> A ∉ {B}
A#5,B          -> A ∉ {5,B}
AB#12        -> A∉{1} OR B∉{2}
AB=12        -> A∈{1} AND B∈{2}

# chains expand into ALL pairwise comparisons.

Examples:
A#B#C
-> A#B
-> A#C
-> B#C

A#B#5
-> A#B
-> A#5
-> B#5

Comma-separated conditions mean AND.

Example:
A#B, A#1 and B#2
means
A#B AND A#1 AND B#2

Check each cell independently.

Return ONLY a JSON object. No explanation. No thinking. No markdown.

Format: {"0": [{"type": "incorrect_parse", "severity": "fail", "detail": "..."}], "1": []}
Keys are cell indices ("0", "1", ...). Values are arrays of issues. Empty array = no issues.
"""


def _strip_markdown_json(raw: str) -> str:
    raw = raw.strip()
    if raw.startswith("```"):
        start = raw.find("\n")
        if start != -1:
            raw = raw[start + 1:]
        if raw.endswith("```"):
            raw = raw[:-3]
        elif raw.endswith("```"):
            raw = raw[:-3]
        raw = raw.strip()
    return raw


def _is_safety_response(text: str) -> bool:
    t = text.strip().lower()
    return not t.startswith("{") and not t.startswith("[") and ("safety" in t or "safe" in t or "content" in t)


def _parse_429_body(resp) -> dict:
    info = {"type": "rate_limited", "reset_ms": None, "detail": "Rate limited"}
    try:
        body = resp.json()
        err = body.get("error", {})
        md = err.get("metadata", {})
        hdrs = md.get("headers", {})
        reset_ms = hdrs.get("X-RateLimit-Reset")
        if reset_ms:
            info["reset_ms"] = int(reset_ms)
        limit = hdrs.get("X-RateLimit-Limit")
        remaining = hdrs.get("X-RateLimit-Remaining")
        msg = err.get("message", "")
        if msg:
            info["detail"] = msg[:200]
        if remaining:
            info["remaining"] = remaining
        if limit:
            info["limit"] = limit
    except Exception:
        info["detail"] = resp.text[:200]
    return info


def _call_llm(prompt: str, model: str, api_key: str) -> tuple[str | None, dict | None]:
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": BATCH_SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0,
    }
    try:
        resp = requests.post(API_BASE, headers=headers, json=payload, timeout=60)
        if resp.status_code == 429:
            return None, _parse_429_body(resp)
        resp.raise_for_status()
        data = resp.json()
        content = data["choices"][0]["message"].get("content")
        if not content:
            return None, {"type": "empty_response", "detail": "Model returned null content"}
        content = content.strip()
        if _is_safety_response(content):
            return None, {"type": "safety_filter", "detail": content[:200]}
        return content, None
    except requests.Timeout:
        return None, {"type": "timeout", "detail": "API request timed out after 60s"}
    except requests.HTTPError as e:
        return None, {"type": "http_error", "detail": f"HTTP {e.response.status_code}"}
    except Exception as e:
        return None, {"type": "exception", "detail": str(e)[:200]}


LLM_ERROR_DIR = Path("llm_errors")


def _save_llm_error(raw: str, batch: list[dict]):
    LLM_ERROR_DIR.mkdir(exist_ok=True)
    ts = time.strftime("%Y%m%d_%H%M%S")
    path = LLM_ERROR_DIR / f"error_{ts}.json"
    payload = {
        "timestamp": ts,
        "raw_response": raw,
        "batch_cells": [
            {
                "cell": b["cell"],
                "template": b["template"],
                "requirements": b["requirements"],
            }
            for b in batch
        ],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"    LLM error saved to {path}")


_FALSE_POSITIVE_PATTERNS = [
    r"trailing comma",
    # r"conflicts with.*and.*which violates",
    # r"can both be",
]


def _is_false_positive(issue: dict) -> bool:
    detail = issue.get("detail", "")
    return any(re.search(p, detail, re.IGNORECASE) for p in _FALSE_POSITIVE_PATTERNS)


def llm_check_batch(batch: list[dict], model: str, api_key: str,
                    api_errors: list[dict]) -> list[list[dict]]:
    lines = []
    batch_cells_info = []
    for i, entry in enumerate(batch):
        cond_strs = " -- ".join(_cond_to_str(c) for c in entry["conditions"])
        template_vars = [ch for ch in entry["template"] if ch.isupper() and ch.isalpha()]
        lines.append(f"[{i}]")
        lines.append(f'  Cell: {entry["cell"]}')
        # lines.append(f'  Vars: {{{",".join(sorted(set(template_vars)))}}}')
        lines.append(f'  Req: {entry["requirements"]}')
        lines.append(f'  "parsed_conditions":{json.dumps([_cond_to_str(c) for c in _flatten_conditions(entry["conditions"])], ensure_ascii=False)}')
        batch_cells_info.append({
            "cell": entry["cell"],
            "template": entry["template"],
            "requirements": entry["requirements"],
        })

    prompt = "Cells to validate:\n\n" + "\n".join(lines)
    print(prompt)

    def try_parse(text: str) -> dict | None:
        cleaned = _strip_markdown_json(text)
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            try:
                decoder = json.JSONDecoder()
                obj, _ = decoder.raw_decode(cleaned)
                if isinstance(obj, dict):
                    return obj
            except (json.JSONDecodeError, ValueError):
                pass
        return None

    max_attempts = 10
    parsed = None
    raw = None
    for attempt in range(max_attempts):
        content, err = _call_llm(prompt, model, api_key)
        if err:
            api_errors.append({
                "attempt": attempt + 1,
                "error": err,
                "batch_cells": batch_cells_info,
            })
            if attempt >= max_attempts - 1:
                break
            reason = err.get("type", "error")
            if reason == "rate_limited":
                reset_ms = err.get("reset_ms")
                if reset_ms:
                    now_ms = int(time.time() * 1000)
                    wait = max(0, (reset_ms - now_ms) // 1000 + 5)
                    print(f"    Rate limit hit — waiting {wait}s until reset...")
                    time.sleep(wait)
                else:
                    print(f"    Rate limit hit, retrying in 10s ({attempt + 2}/{max_attempts})...")
                    time.sleep(10)
            else:
                delay = min(2 ** attempt * 5, 30)
                print(f"    {reason}, retrying in {delay}s ({attempt + 2}/{max_attempts})...")
                time.sleep(delay)
            continue

        raw = content
        parsed = try_parse(raw)
        if parsed is not None:
            if isinstance(parsed, list):
                parsed = {str(i): v for i, v in enumerate(parsed)}
            break

        _save_llm_error(raw, batch)
        if attempt < max_attempts - 1:
            print(f"    Parse error, retrying ({attempt + 2}/{max_attempts})...")
            time.sleep(1)

    if parsed is None:
        _save_llm_error(raw, batch)
        return [[{"stage": "llm", "type": "api_error", "severity": "warn",
                   "detail": "LLM API call failed after all retries"}] for _ in batch]

    results = []
    if isinstance(parsed, dict) and "type" in parsed:
        parsed = {"0": [parsed]}
    for i in range(len(batch)):
        cell_issues = parsed.get(str(i), [])
        if isinstance(cell_issues, list):
            filtered = [issue for issue in cell_issues if not _is_false_positive(issue)]
            for issue in filtered:
                issue["stage"] = "llm"
            results.append(filtered)
        else:
            results.append([])
    return results


def _log_failure(result: dict):
    if result["status"] == "PASS":
        return
    entry = {
        "row": result["row"],
        "status": result["status"],
        "cell": result["cell"],
        "template": result.get("template", ""),
        "issues": [{"stage": i["stage"], "type": i["type"], "severity": i["severity"], "detail": i["detail"][:120]} for i in result["issues"]],
    }
    with open(FAIL_LOG, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def _status(issues: list[dict]) -> str:
    if any(i.get("severity") == "fail" for i in issues):
        return "FAIL"
    if issues:
        return "WARN"
    return "PASS"


def _print_table(results: list[dict]):
    header = f"{'Row':>6} | {'Status':6} | {'Template':30} | Issues"
    sep = "-" * len(header)
    print(header)
    print(sep)
    for r in results:
        issues_str = "; ".join(f"[{i['type']}] {i['detail'][:60]}" for i in r["issues"])
        tmpl = (r.get("template") or "")[:30]
        print(f"{r['row']:>6} | {r['status']:6} | {tmpl:30} | {issues_str}")
    print(sep)

    pass_count = sum(1 for r in results if r["status"] == "PASS")
    warn_count = sum(1 for r in results if r["status"] == "WARN")
    fail_count = sum(1 for r in results if r["status"] == "FAIL")
    print(f"Total: {len(results)} | PASS: {pass_count} | WARN: {warn_count} | FAIL: {fail_count}")


def _write_json(results: list[dict], api_errors: list[dict], path: str, model: str = ""):
    report = {
        "metadata": {
            "model": model,
            "provider": _detect_provider(),
            "total_rows": len(results),
            "date": time.strftime("%Y-%m-%d %H:%M:%S"),
        },
        "summary": {
            "pass": sum(1 for r in results if r["status"] == "PASS"),
            "warn": sum(1 for r in results if r["status"] == "WARN"),
            "fail": sum(1 for r in results if r["status"] == "FAIL"),
        },
        "failures": [{"row": r["row"], "cell": r["cell"], "template": r["template"],
                       "issues": r["issues"]} for r in results if r["status"] == "FAIL"],
        "warnings": [{"row": r["row"], "cell": r["cell"], "template": r["template"],
                       "issues": r["issues"]} for r in results if r["status"] == "WARN"],
        "api_errors": api_errors,
        "results": results,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


def main():
    parser = argparse.ArgumentParser(description="Validate cell formats and LLM-check parser compatibility.")
    parser.add_argument("--max-rows", type=int, default=None, help="Limit rows to process")
    parser.add_argument("--index", type=int, default=None, help="Start processing from this row number")
    parser.add_argument("--model", default=None, help=f"Model (default: GROQ_MODEL/OPENROUTER_MODEL env, provider={PROVIDER})")
    parser.add_argument("--output", default="validation_report.json", help="JSON output path")
    args = parser.parse_args()

    api_key = os.getenv("GROQ_API_KEY") or os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        print("ERROR: neither GROQ_API_KEY nor OPENROUTER_API_KEY is set", file=sys.stderr)
        sys.exit(1)

    provider = _detect_provider()
    if provider == "groq":
        model = args.model or os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
    else:
        model = args.model or os.getenv("OPENROUTER_MODEL", "openrouter/free")

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    results = []
    api_errors: list[dict] = []
    total_rows = ws.max_row
    start_row = args.index or 2
    end_row = total_rows if args.max_rows is None else min(start_row + args.max_rows - 1, total_rows)

    pending = []
    pending_cell_idx = []

    def flush_batch():
        if not pending:
            return
        batch_results = llm_check_batch(pending, model, api_key, api_errors)
        for idx, extra_issues in zip(pending_cell_idx, batch_results):
            results[idx]["issues"].extend(extra_issues)
            results[idx]["status"] = _status(results[idx]["issues"])
            r = results[idx]
            _log_failure(r)
            print(f"  Row {r['row']:4d} | {r['status']:6s} | issues: {len(r['issues'])}")
        pending.clear()
        pending_cell_idx.clear()
        time.sleep(BATCH_DELAY)

    for row_idx in range(start_row, end_row + 1):
        cell_value = ws.cell(row=row_idx, column=2).value
        if cell_value is None:
            continue

        cell_value = str(cell_value).strip()
        if not cell_value:
            continue

        issues = []

        parsed = parse_dinh_dang(cell_value)
        issues.extend(check_format(cell_value, parsed))

        requirements = parsed.get("requirements")
        conditions = []
        if requirements:
            conditions = parse(requirements)
            issues.extend(check_parse(requirements, conditions))
            issues.extend(check_variables(parsed.get("number", ""), conditions))

        result = {
            "row": row_idx,
            "cell": cell_value,
            "template": parsed.get("number", ""),
            "requirements": requirements,
            "parsed_conditions": " -- ".join(_cond_to_str(c) for c in conditions) if conditions else "",
            "status": _status(issues),
            "issues": issues,
        }
        idx = len(results)
        results.append(result)

        if requirements:
            pending.append({
                "cell": cell_value,
                "template": parsed.get("number", ""),
                "requirements": requirements,
                "conditions": conditions,
            })
            pending_cell_idx.append(idx)

            if len(pending) >= BATCH_SIZE:
                flush_batch()
        else:
            _log_failure(result)
            print(f"  Row {row_idx:4d} | {result['status']:6s} | (no rules)")

    if pending:
        flush_batch()

    wb.close()

    print()
    _print_table(results)
    _write_json(results, api_errors, args.output, model)
    print(f"\nReport written to {args.output}")


if __name__ == "__main__":
    main()
